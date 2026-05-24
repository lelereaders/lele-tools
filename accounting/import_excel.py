#!/usr/bin/env python3
"""
One-time import of 樂樂文化 Excel ledgers into Google Sheets via Apps Script.

Usage (PowerShell):
  cd accounting
  $env:GAS_URL   = "https://script.google.com/macros/s/YOUR_ID/exec"
  $env:GAS_TOKEN = "your-token"
  python import_excel.py
"""
import os
import sys
import datetime
import requests
from pathlib import Path
import openpyxl

GAS_URL   = os.environ["GAS_URL"]
GAS_TOKEN = os.environ["GAS_TOKEN"]
EXCEL_DIR = Path(__file__).parent.parent / "樂樂文化-記帳"
SHEET_NAME = "日記帳(每日記錄)"
SKIP_CATEGORIES = {"Carryforward 前期結轉"}
BATCH_SIZE = 400


def parse_date(value) -> str:
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    s = str(value) if value is not None else ""
    return s[:10] if len(s) >= 10 else ""


def load_excel(path: Path) -> list:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR opening {path.name}: {e}", file=sys.stderr)
        return []
    if SHEET_NAME not in wb.sheetnames:
        print(f"  skip (no sheet '{SHEET_NAME}'): {path.name}")
        return []
    ws = wb[SHEET_NAME]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        tid = row[0] if len(row) > 0 else None
        if not tid or not str(tid).startswith("T"):
            continue
        category = str(row[3]) if len(row) > 3 and row[3] is not None else ""
        if category in SKIP_CATEGORIES:
            continue
        amount_raw = row[5] if len(row) > 5 else None
        amount = float(amount_raw) if amount_raw is not None else 0.0
        twd_raw = row[9] if len(row) > 9 else None
        twd_amount = float(twd_raw) if twd_raw is not None else amount
        rows.append({
            "id":           str(tid),
            "date":         parse_date(row[1] if len(row) > 1 else None),
            "description":  str(row[2]) if len(row) > 2 and row[2] is not None else "",
            "category":     category,
            "currency":     str(row[4]) if len(row) > 4 and row[4] is not None else "TWD",
            "amount":       amount,
            "from_account": str(row[6]) if len(row) > 6 and row[6] is not None else "",
            "to_account":   str(row[7]) if len(row) > 7 and row[7] is not None else "",
            "notes":        str(row[8]) if len(row) > 8 and row[8] is not None else "",
            "twd_amount":   twd_amount,
            "created_at":   datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.000Z"),
            "source":       "excel_import",
        })
    wb.close()
    return rows


def push_batch(batch: list) -> None:
    resp = requests.post(
        GAS_URL,
        json={"token": GAS_TOKEN, "action": "addTransactionBatch", "data": batch},
        timeout=120,
    )
    resp.raise_for_status()
    result = resp.json()
    if "error" in result:
        raise RuntimeError(f"GAS error: {result['error']}")


def main():
    files = sorted(EXCEL_DIR.glob("樂樂-雲端帳簿-*.xlsx"))
    if not files:
        print(f"No Excel files found in {EXCEL_DIR}")
        sys.exit(1)

    all_rows: list = []
    seen_ids: set = set()
    for f in files:
        print(f"Reading {f.name}...")
        rows = load_excel(f)
        before = len(all_rows)
        for r in rows:
            if r["id"] not in seen_ids:
                seen_ids.add(r["id"])
                all_rows.append(r)
        added = len(all_rows) - before
        print(f"  {len(rows)} rows read, {added} new unique")

    print(f"\nTotal unique transactions: {len(all_rows)}")
    if len(all_rows) == 0:
        print("Nothing to import.")
        sys.exit(0)

    print("Pushing to Google Sheets in batches...")
    for i in range(0, len(all_rows), BATCH_SIZE):
        batch = all_rows[i : i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        print(f"  Batch {batch_num}: {len(batch)} rows...", end=" ", flush=True)
        push_batch(batch)
        print("ok")

    print("\nImport complete.")


if __name__ == "__main__":
    main()
